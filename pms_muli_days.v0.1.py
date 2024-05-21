import numpy as np
import pandas as pd
import openpyxl
import calendar
import math
from datetime import datetime, timedelta
from chinese_calendar import is_workday
from openpyxl.styles import PatternFill, Border, Side
from xls2xlsx import XLS2XLSX


# 在原基础上适用于任何月份/周，自定义目录，文件名手动输入，自动将xls转为xlsx文件
# 时长统计符合部门的规则，开始时间向前取整，时长向上取整，比如从9：50开始到10：10结束，命中2个小时
# 每天计算时长按10小时算，以贴合部门KPI
# 工作日使用chinese_calendar判断，去除法定节假日，计算法定调休日，与公司考勤一致
# 不再需要修改文件路径及工作日

# 定义Excel文件路径
pms_file_path = "E:/WORK/Python_PMS/"  # 修改为实际文件目录

# 定义人员清单文件名，供后续处理合并参考
file_path_list = f"{pms_file_path}KPI-List.xlsx"

# 定义输入xls文件名
filename = input("输入待处理的文件名:")
file_path_xls = f"{pms_file_path}{filename}.xls"
# 定义输入起止时间
input_date_start_str = input("请输入开始日期 (格式为 YYYY/M/D): ")
input_date_end_str = input("请输入结束日期 (格式为 YYYY/M/D): ")
input_date_start = datetime.strptime(input_date_start_str, '%Y/%m/%d').date()
input_date_end =datetime.strptime(input_date_end_str, '%Y/%m/%d').date()
#input_date_start = input_date_start.date()
#input_date_end = input_date_end.date()

print(f"输入开始日期为：{input_date_start}")
print(f"输入结束日期为：{input_date_end}")
# 判断xls文件中日志所属年份及月份
xls_df = pd.read_excel(file_path_xls)
# 获取A2表格中的日期，格式为YYYY-MM-DD
xls_date_end = datetime.strptime(str(xls_df.iloc[1, 0]), '%Y-%m-%d %H:%M:%S').date()
xls_date_start = datetime.strptime(str(xls_df.iloc[-1, 0]), '%Y-%m-%d %H:%M:%S').date()
print(f"日志数据开始日期为：{xls_date_start}")
print(f"日志数据结束日期为：{xls_date_end}")
if xls_date_start < input_date_start or xls_date_end > input_date_end:
    print(f"！！！！！！数据所属的日期与指定不符，可能不准，请核对！！！！！！")

pms_year = xls_date_end.year
pms_month = xls_date_end.month
pms_month_start = xls_date_start.month

workdays = 0
# 计算工作日
total_days = (input_date_end - input_date_start).days + 1
print(f"两个输入日期之间相隔{total_days}天")
workdays = 0
workday_count = input_date_start
for day in range(1, total_days + 1):
    if is_workday(workday_count):
        workdays += 1
    workday_count += timedelta(days=1)
print(f"工作日一共:{workdays}天")

# 定义KPI考核项目日志时长基数#####按需修改，80%为满分
kpi_base = 0.8
# 定义KPI考核每日应当日志时长#####按需修改
work_hours_per_day = 10

# 定义后续Excel文件名
file_path = f"{pms_file_path}{input_date_start}-{input_date_end}.xlsx"
file_path_output = f"{pms_file_path}{input_date_start}-{input_date_end}-output.xlsx"

# xls转为xlsx
x2x = XLS2XLSX(file_path_xls)
x2x.to_xlsx(file_path)

# ##############################初始xls判断及重新按年月输出为xlsx文件结束#####################################
# #####################################进行时长及日志类型处理###############################################


# 使用 openpyxl 打开 Excel 文件
workbook = openpyxl.load_workbook(file_path)
sheet = workbook.active  # 默认选择活动工作表

# 读取Excel文件project_name is None and
for row in range(2, sheet.max_row + 1):
    project_name = str(sheet.cell(row=row, column=5).value).strip()
    work_content = str(sheet.cell(row=row, column=6).value).strip()
    if (project_name is None or project_name.strip() == "") and("假" in work_content or "休" in work_content):
        # 如果在工作内容中找到 "假" 或 "休" 关键字，则在E列中写入 "请假"
        sheet.cell(row=row, column=5, value="请假")
    elif (project_name is not None and sheet.cell(row=row, column=5).value is None) and ("假" in work_content or "休" in work_content):
        # 如果在工作内容中找到 "假" 或 "休" 关键字，则在E列中写入 "请假"
        sheet.cell(row=row, column=5, value="请假")


# 遍历每一行，搜索 "工作内容" 列中的关键字
for row in range(2, sheet.max_row + 1):
    project_name = str(sheet.cell(row=row, column=5).value)
    if project_name is None or project_name.strip() == "":
        # 如果项目名称为空，则在F列中写入 "日常"
        sheet.cell(row=row, column=5, value="日常")
    elif project_name is not None and sheet.cell(row=row, column=5).value is None:
        # 如果项目名称不为空且E列为空，则在E列中写入 "日常"
        sheet.cell(row=row, column=5, value="日常")

# ########################工程师名称规范化修改################
sheet.cell(row=1, column=4, value="姓名")
for row in range(2, sheet.max_row + 1):
    name_fix = str(sheet.cell(row=row, column=4).value)
    if name_fix.strip() == "陈帅(武汉)":
        sheet.cell(row=row, column=4, value="陈帅")
# ########################################################

# 保存更新后的数据到xlsx文件
workbook.save(file_path)

# 打开工作簿
workbook = openpyxl.load_workbook(file_path)  # 需要需改路径######################################

# 选择工作表
sheet = workbook.active

# 插入新列，将"时长"列插入到第五列
sheet.insert_cols(5)

# 设置"时长"列的表头
sheet.cell(row=1, column=5, value="时长")

# 遍历每一行，计算时长并写入第5列
for row in range(2, sheet.max_row + 1):
    start_time_str = sheet.cell(row=row, column=2).value
    end_time_str = sheet.cell(row=row, column=3).value
    leave_str = sheet.cell(row=row, column=6).value
    if start_time_str is not None and end_time_str is not None:
        # 将时间字符串解析为时间对象，并将开始时间向前取整
        start_time = datetime.strptime(start_time_str, '%H:%M').replace(minute=0)
        end_time = datetime.strptime(end_time_str, '%H:%M')
        # 计算时长
        duration = (end_time - start_time).total_seconds() / 3600
        # 使用math.ceil向上取整
        duration = math.ceil(duration)
        # 将时长写入第五列，全天请假最多算8小时
        if leave_str == '请假' and duration > 8:
            duration = 8
        sheet.cell(row=row, column=5, value=duration)

# 保存工作簿
workbook.save(file_path)
# 关闭工作簿
workbook.close()
# #####################################完成时长及日志类型处理###############################################


# 读取Excel文件
df = pd.read_excel(file_path)

# 根据 "项目名称" 列筛选出工作记录
work_df = df[df['项目名称'] != '请假'].copy()  # 使用.copy()创建一个切片的副本

# A.计算日志时长
work_df['时长'] = work_df['时长'].abs()  # 将负数时长转换为正数
# 判断数据月份是否全部属于当前月，进行汇总
total_work_duration = work_df.groupby(['姓名'])['时长'].sum().reset_index()
# 创建工作日的列表，按月份匹配天数

# ##############合并KPI_List开始#######
# 读取List文件，并指定工号为字符串类型
list_df = pd.read_excel(file_path_list, dtype={'工号': str})
# 合并List
total_work_duration = list_df.merge(total_work_duration, on=['姓名'], how='left')
# ##############合并KPI_List完成#######

# 计算工作日时长
total_work_duration['工作日'] = workdays
total_work_duration['工作日时长'] = total_work_duration['工作日'] * work_hours_per_day

# B.计算请假时长
leave_df = df[df['项目名称'] == '请假'].copy()  # 使用.copy()创建一个切片的副本
leave_df['时长'] = leave_df['时长'].abs()  # 将负数时长转换为正数
# 判断数据月份是否全部属于当前月，进行汇总
total_leave_duration = leave_df.groupby(['姓名'])['时长'].sum().reset_index()
# C.合并总日志时长=日志时长+请假时长+居家办公+日常时长
result_df = total_work_duration.merge(total_leave_duration, on=['姓名'], how='left')
# D.合并居家办公加日常=居家办公+日常时长
home_office_df = df[df['项目名称'] == '居家办公'].copy()  # 使用.copy()创建一个切片的副本
daily_df = df[df['项目名称'] == '日常'].copy()  # 使用.copy()创建一个切片的副本
# 组合居家办公和日常
leave_and_home_office_df = pd.concat([home_office_df, daily_df])
leave_and_home_office_df['时长'] = leave_and_home_office_df['时长'].abs()  # 将负数时长转换为正数
# 判断数据月份是否全部属于当前月，进行汇总
leave_and_home_office_duration = leave_and_home_office_df.groupby(['姓名'])['时长'].sum().reset_index()
leave_and_home_office_duration.rename(columns={'时长': '居家办公加日常'}, inplace=True)
result_df = result_df.merge(leave_and_home_office_duration, on=['姓名'], how='left')

result_df.fillna(0, inplace=True)  # 填充NaN值为0，表示没有请假时长


# 重命名列名
result_df.rename(columns={'时长_x': '日志时长', '时长_y': '请假时长', '居家办公加日常': '日常日志时长'}, inplace=True)
result_df['日志时长'] = result_df['日志时长'] + result_df['请假时长']
# 计算项目日志时长
result_df['项目日志时长'] = result_df['日志时长'] - result_df['请假时长'] - result_df['日常日志时长']
# 计算项目日志占比，如果
result_df['项目日志占比'] = result_df['项目日志时长'] / result_df['工作日时长']

# 计算每个支持工程师的排名，根据项目日志占比排名
result_df['排名'] = result_df.groupby('姓名')['项目日志占比'].rank(ascending=False, method='min')
# 保留小数点后两位
result_df = result_df.round(2)


# 根据项目日志时长排名
result_df['排名'] = result_df['项目日志时长'].rank(ascending=False, method='min')

# 将排名按项目日志占比升序排序
result_df.sort_values(by='排名', ascending=True, inplace=True)

# 重新计算总支持工程师人数
total_support_engineers = result_df['姓名'].nunique()

# 更新工程师的分类
result_df['项目日志时长'] = result_df['项目日志时长'].rank(ascending=False, method='min')

# 根据项目日志排名将工程师分类
result_df['排名'] = result_df['项目日志时长'].apply(
    lambda rank: f"超过{int(((total_support_engineers - rank) / total_support_engineers) * 10) * 10}%" if rank > 0 else "后十名"
)

# 找到排名超过0%的工程师并将其命名为"后十名"
result_df.loc[result_df['排名'] == '超过0%', '排名'] = '后十名'

# 保留小数点后两位
result_df = result_df.round(2)

# 添加项目日志时长、项目日志占比这三列
result_df['项目日志时长'] = result_df['项目日志时长'] = result_df['日志时长'] - result_df['请假时长'] - result_df['日常日志时长']
result_df['项目日志占比'] = result_df['项目日志时长'] / result_df['工作日时长']

# 添加KPI
result_df['KPI参考'] = result_df['项目日志占比'] / kpi_base
# 重置索引
result_df.reset_index(drop=True, inplace=True)
result_df['项目日志占比'] = result_df['项目日志占比'].round(2)
result_df['KPI参考'] = result_df['KPI参考'].round(2)
result_df['KPI有效值（0-150）'] = result_df['KPI参考'].clip(upper=1.5)

# 重新排列列的顺序
result_df['日志区间'] = f"{input_date_start}至{input_date_end}"
new_column_order = ['姓名', '工号', 'Base地', '岗位类别', '是否外包', '日志区间', '工作日', '工作日时长', '日志时长', '请假时长', '日常日志时长', '项目日志时长', '项目日志占比', 'KPI参考', '排名', 'KPI有效值（0-150）', '邮箱', '备注']
# 选择并重新排列列
result_df = result_df[new_column_order]
# 将结果保存到Excel文件
result_df.to_excel(file_path_output, index=False)

# 使用 openpyxl 打开 Excel 文件
workbook = openpyxl.load_workbook(file_path_output)
sheet = workbook.active  # 默认选择活动工作表
border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
red_fill = PatternFill(start_color='ffc7ce', end_color='ffc7ce', fill_type='solid')
green_fill = PatternFill(start_color='ceffc7', end_color='ceffc7', fill_type='solid')
for row in sheet.iter_rows():
    for cell in row:
        cell.border = border

for row in range(2, sheet.max_row + 1):
    sheet.cell(row=row, column=13).number_format = '0%'
    sheet.cell(row=row, column=14).number_format = '0%'
    if sheet.cell(row=row, column=14).value < 1:
        sheet.cell(row=row, column=14).fill = red_fill
    sheet.cell(row=row, column=16).number_format = '0%'
    if sheet.cell(row=row, column=16).value < 1:
        sheet.cell(row=row, column=16).fill = red_fill
    elif sheet.cell(row=row, column=16).value >= 1.25:
        sheet.cell(row=row, column=16).fill = green_fill



# 保存工作簿
workbook.save(file_path_output)
# 关闭工作簿
workbook.close()
