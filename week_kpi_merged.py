import os
import pandas as pd
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill, Border, Side
# 合并各小组的周KPI为区域周KPI

# 定义Excel文件路径
pms_file_path = "E:/WORK/Python_PMS/"
# 定义人员清单文件名，供后续处理合并参考
file_path_list = f"{pms_file_path}KPI-List.xlsx"
# 设置包含Excel文件的目录路径
week_kpi_path = "E:/WORK/Python_PMS/区域周KPI合并/"
file_path_output = f"{week_kpi_path}周kpi合并_{datetime.now().strftime("%Y-%m-%d")}.xlsx"
# pms_file_path = "E:/WORK/Python_PMS/"
# file_path_output = f"{pms_file_path}周kpi合并_{datetime.now().strftime("%Y-%m-%d")}.xlsx"

# 初始化一个空的DataFrame，用于存储合并后的数据
kpi_merged_df = pd.DataFrame()

# 遍历目录中的所有文件
for filename in os.listdir(week_kpi_path):
    # 检查文件扩展名是否为xlsx
    if filename.endswith('.xlsx') and not filename.startswith('周kpi合并'):
        # 构建完整的文件路径
        file_path = os.path.join(week_kpi_path, filename)
        # 读取Excel文件
        kpi_df = pd.read_excel(file_path)
        # 填充空白单元格
        # kpi_df.fillna(0, inplace=True)
        # 将读取的数据追加到kpi_merged_df中
        kpi_merged_df = pd.concat([kpi_merged_df, kpi_df], ignore_index=True)

# 保存合并后的DataFrame到一个新的Excel文件
# 读取List文件，并指定工号为字符串类型
list_df = pd.read_excel(file_path_list, sheet_name='人员信息', dtype={'工号': str})

# 合并List
kpi_merged_df = kpi_merged_df.merge(list_df[['姓名', '备注']], on=['姓名'], how='left')
# 写入文件
kpi_merged_df.to_excel(file_path_output, index=False)


# 使用 openpyxl 打开 Excel 文件
workbook = openpyxl.load_workbook(file_path_output)
sheet = workbook.active  # 默认选择活动工作表
border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
red_fill = PatternFill(start_color='ffc7ce', end_color='ffc7ce', fill_type='solid')
green_fill = PatternFill(start_color='ceffc7', end_color='ceffc7', fill_type='solid')
for row in sheet.iter_rows():
    for cell in row:
        cell.border = border

for row in range(2, sheet.max_row + 1):
    if sheet.cell(row=row, column=6).value < 0:
        sheet.cell(row=row, column=6).fill = red_fill
    elif sheet.cell(row=row, column=6).value > 3:
        sheet.cell(row=row, column=6).fill = green_fill

# 保存工作簿
workbook.save(file_path_output)
# 关闭工作簿
workbook.close()

print(f'All Excel files have been merged into {file_path_output}')