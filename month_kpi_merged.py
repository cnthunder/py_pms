import os
import pandas as pd
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill, Border, Side
# 合并区域周KPI为区域月KPI

# 定义Excel文件路径
pms_file_path = "E:/WORK/Python_PMS/"
# 定义人员清单文件名，供后续处理合并参考
file_path_list = f"{pms_file_path}KPI-List.xlsx"
# 设置包含Excel文件的目录路径
month_kpi_path = "E:/WORK/Python_PMS/区域月KPI合并/"
file_path_output = f"{month_kpi_path}/合并后/月kpi合并_{datetime.now().strftime("%Y-%m-%d")}.xlsx"

# 初始化一个空的DataFrame，用于存储合并后的数据
kpi_merged_df = pd.DataFrame()
week_count = 0
# 遍历目录中的所有文件
for filename in os.listdir(month_kpi_path):
    # 检查文件扩展名是否为xlsx
    if filename.endswith('.xlsx') and filename.startswith('周kpi合并'):
        # 构建完整的文件路径
        file_path = os.path.join(month_kpi_path, filename)
        week_count += 1
        print(f'输入周KPI表格{week_count}：{file_path}')
        # 读取Excel文件
        kpi_df = pd.read_excel(file_path)
        kpi_df.drop(['项目日志占比', '备注_x', '备注_y'], axis=1, inplace=True)
        kpi_df.rename(columns={'周报准确性（评分项）': '周报准确性'}, inplace=True)
        # 填充空白单元格
        # kpi_df.fillna(0, inplace=True)
        # 将读取的数据追加到kpi_merged_df中
        kpi_merged_df = pd.concat([kpi_merged_df, kpi_df], ignore_index=True)
        # 计算周数


# 去掉备注列
#kpi_merged_df.drop(['备注_x', '备注_y'], axis=1, inplace=True)
# 保存合并后的DataFrame到一个新的Excel文件
# 读取List文件，并指定工号为字符串类型
list_df = pd.read_excel(file_path_list, sheet_name='人员信息', dtype={'工号': str})

# 合并List
#kpi_merged_df = kpi_merged_df.merge(list_df[['姓名', '备注']], on=['姓名'], how='left')
kpi_merged_df.fillna(0, inplace=True)
# 合并计算扣分项
kpi_merged_plus_df = kpi_merged_df.groupby(['姓名'])[['日志及时性', '日志准确性', '周报及时性', '工作报备及时性', '工作报备准确性', '工作反馈及时性', '工作反馈准确性']].sum().reset_index()
# 合并计算评分项
kpi_merged_avg_df = (kpi_merged_df.groupby(['姓名'])['周报准确性'].sum()/week_count).reset_index()
month_kpi_df = kpi_merged_plus_df.merge(kpi_merged_avg_df, on=['姓名'], how='left')
# 读取月KPI有效值
kpi_int_df = pd.DataFrame()
for filename in os.listdir(month_kpi_path):
    # 检查文件扩展名是否为xlsx
    if filename.endswith('-output.xlsx') and filename.startswith('2024-'):
        # 构建完整的文件路径
        file_path = os.path.join(month_kpi_path, filename)
        print(f'项目日志占比参考:{file_path}')
        # 读取Excel文件
        kpi_int_df = pd.read_excel(file_path)
# 合并项目日志占比项，将月日志KPI有效值作为项目日志占比参考
month_kpi_df = month_kpi_df.merge(kpi_int_df[['姓名', '日志区间', 'KPI有效值（0-150）']], on=['姓名'], how='left')
month_kpi_df.rename(columns={'KPI有效值（0-150）': '项目日志占比', '日志区间': '月份'}, inplace=True)

# 从KPI_List文件里导入人员信息
month_kpi_df = month_kpi_df.merge(list_df[['姓名', '工号', 'Base地', '岗位类别', '外包项目', '邮箱', '备注']], on=['姓名'], how='left')
# 增加月KPI参考项
month_kpi_df['月KPI参考'] = ''
# 根据各项权重计算月KPI，并取整
month_kpi_df['月KPI参考'] = (100-abs(month_kpi_df['日志及时性']))/100*10 + (100-abs(month_kpi_df['日志准确性']))/100*10 + month_kpi_df['项目日志占比']*40 + (100-abs(month_kpi_df['周报及时性']))/100*10 + (100-abs(month_kpi_df['周报准确性']))/100*10 + (100-abs(month_kpi_df['工作报备及时性']))/100*5 + (100-abs(month_kpi_df['工作报备准确性']))/100*5 + (100-abs(month_kpi_df['工作反馈及时性']))/100*5 + (100-abs(month_kpi_df['工作反馈准确性']))/100*5
month_kpi_df['月KPI参考'] = month_kpi_df['月KPI参考'].round(2)
# 按月KPI参考进行排序
month_kpi_df.sort_values(by='月KPI参考', ascending=False, inplace=True)

# 根据月KPI计算排名
month_kpi_df['排名'] = month_kpi_df.groupby('姓名')['月KPI参考'].rank(ascending=False, method='min')
month_kpi_df['排名'] = month_kpi_df['月KPI参考'].rank(ascending=False, method='min')
# 计算总支持人数
mens_count = month_kpi_df['姓名'].nunique()
# 排名计算
month_kpi_df['排名计算'] = month_kpi_df['月KPI参考'].rank(ascending=False, method='min')
# 根据月KPI参考排名将工程师分类
month_kpi_df['排名'] = month_kpi_df['排名计算'].apply(
    lambda rank: f"超过{int(((mens_count - rank) / mens_count) * 10) * 10}%" if rank > 0 else "后十名"
)
# 找到排名超过0%的工程师并将其命名为"后十名"
month_kpi_df.loc[month_kpi_df['排名'] == '超过0%', '排名'] = '后十名'

# 重新排序，过滤输出列
new_column_order = ['姓名',  '工号', 'Base地', '岗位类别', '外包项目', '月份',
                    '日志及时性', '日志准确性', '项目日志占比', '周报及时性', '周报准确性',
                    '工作报备及时性', '工作报备准确性', '工作反馈及时性', '工作反馈准确性',
                    '月KPI参考', '排名', '邮箱', '备注']
month_kpi_df = month_kpi_df[new_column_order]

# 写入输出表格
month_kpi_df.to_excel(file_path_output, index=False)

# 使用 openpyxl 打开 Excel 文件
workbook = openpyxl.load_workbook(file_path_output)
sheet = workbook.active  # 默认选择活动工作表
border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
red_fill = PatternFill(start_color='ffc7ce', end_color='ffc7ce', fill_type='solid')
green_fill = PatternFill(start_color='ceffc7', end_color='ceffc7', fill_type='solid')
for row in sheet.iter_rows():
    for cell in row:
        cell.border = border

for row in range(2, sheet.max_row + 1):
    sheet.cell(row=row, column=9).number_format = '0%'
    if sheet.cell(row=row, column=16).value < 95:
        sheet.cell(row=row, column=16).fill = red_fill
    elif sheet.cell(row=row, column=16).value > 110:
        sheet.cell(row=row, column=16).fill = green_fill

# 保存工作簿
workbook.save(file_path_output)
# 关闭工作簿
workbook.close()

print(f'All Excel files have been merged into {file_path_output}')